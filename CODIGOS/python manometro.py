"""
╔══════════════════════════════════════════════════════════════╗
║         ESP32 · MQ-135 Dashboard Profesional                ║
║         Requiere: pip install pyserial openpyxl              ║
║         Uso: python mq135_dashboard.py                       ║
║         (Muestra datos en tiempo real)                       ║
║         (Excel guarda el promedio cada 10 segundos)          ║
╚══════════════════════════════════════════════════════════════╝
"""

import tkinter as tk
from tkinter import ttk
import math
import serial
import serial.tools.list_ports
import threading
import time
import collections
import random
import os
import csv
import datetime
import tempfile

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
#  CONFIGURACIÓN
# ──────────────────────────────────────────────
DEFAULT_PORT  = "COM3"
DEFAULT_BAUD  = 115200
ADC_MAX       = 4095
VREF          = 3.3
HISTORY_LEN   = 80
UPDATE_MS     = 50
CSV_UPDATE_EVERY = 10   # flush del buffer CSV cada N lecturas

# Rangos de calidad del aire (ADC 12-bit)
AIR_QUALITY = [
    (0,    1200, "EXCELENTE",  "#00ff88"),
    (1200, 2000, "BUENO",      "#aaff00"),
    (2000, 2700, "MODERADO",   "#ff9900"),
    (2700, 3400, "MALO",       "#ff6600"),
    (3400, 4095, "PELIGROSO",  "#ff3333"),
]

# Paleta dark HUD
C = {
    "bg":      "#050c12",
    "panel":   "#0a1520",
    "border":  "#1a3a5a",
    "text":    "#8bbbbb",
    "bright":  "#cceeee",
    "hud":     "#00ffff",
    "ok":      "#00ff88",
    "warn":    "#ff9900",
    "danger":  "#ff3333",
    "amber":   "#ffaa00",
}

# ──────────────────────────────────────────────
#  UTILIDADES
# ──────────────────────────────────────────────
def classify(val):
    for lo, hi, label, color in AIR_QUALITY:
        if lo <= val <= hi:
            return label, color
    return "DESCONOCIDO", C["text"]

def needle_coords(cx, cy, r, angle_deg):
    rad = math.radians(angle_deg)
    x = cx + r * math.cos(rad)
    y = cy - r * math.sin(rad)
    return x, y

def adc_to_angle(val):
    return 180.0 - (val / ADC_MAX * 180.0)


# ══════════════════════════════════════════════
#  CSV LOGGER (registra cada lectura)
# ══════════════════════════════════════════════
class CsvLogger:
    HEADERS = ["N°", "Fecha", "Hora", "ADC", "Voltaje_V",
               "Escala_%", "Calidad", "Sesion"]

    def __init__(self, filepath: str):
        self.filepath = filepath
        self._lock    = threading.Lock()
        self._count   = 0
        self._buffer  = []
        with open(self.filepath, "w", newline="", encoding="utf-8-sig") as f:
            csv.writer(f).writerow(self.HEADERS)

    def log(self, val: int, label: str, session_name: str):
        dt      = datetime.datetime.now()
        voltage = round(val * VREF / ADC_MAX, 4)
        pct     = round(val / ADC_MAX * 100, 2)
        row = [
            self._count + 1,
            dt.strftime("%Y-%m-%d"),
            dt.strftime("%H:%M:%S.") + f"{dt.microsecond // 1000:03d}",
            val, voltage, pct, label, session_name,
        ]
        with self._lock:
            self._count += 1
            self._buffer.append(row)
            if self._count % CSV_UPDATE_EVERY == 0:
                self._flush()

    def _flush(self):
        try:
            with open(self.filepath, "a", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                for r in self._buffer:
                    w.writerow(r)
            self._buffer.clear()
        except Exception:
            pass

    def close(self):
        with self._lock:
            self._flush()


# ══════════════════════════════════════════════
#  EXCEL LOGGER (versión robusta con guardado atómico)
# ══════════════════════════════════════════════
class ExcelLogger:
    COLOR_HEADER_BG   = "0D1F2D"
    COLOR_HEADER_FG   = "00FFFF"
    COLOR_SUBHDR_BG   = "1A3A5A"
    COLOR_SUBHDR_FG   = "CCEEEE"
    COLOR_ROW_EVEN    = "0A1520"
    COLOR_ROW_ODD     = "0D1F30"
    COLOR_EXCELENTE   = "00C86A"
    COLOR_BUENO       = "88CC00"
    COLOR_MODERADO    = "FF8800"
    COLOR_MALO        = "FF5500"
    COLOR_PELIGROSO   = "FF2222"
    COLOR_STATS_BG    = "122030"

    CALIDAD_COLORS = {
        "EXCELENTE": COLOR_EXCELENTE,
        "BUENO":     COLOR_BUENO,
        "MODERADO":  COLOR_MODERADO,
        "MALO":      COLOR_MALO,
        "PELIGROSO": COLOR_PELIGROSO,
    }

    def __init__(self, filepath: str, logger_callback=None):
        self.filepath   = filepath
        self.wb         = Workbook()
        self._row       = 2
        self._count     = 0
        self._lock      = threading.Lock()
        self.log        = logger_callback or print

        self.ws_data  = self.wb.active
        self.ws_data.title = "Lecturas"
        self.ws_stats = self.wb.create_sheet("Estadísticas")
        self.ws_info  = self.wb.create_sheet("Info")

        self._setup_data_sheet()
        self._setup_stats_sheet()
        self._setup_info_sheet()
        self._save(initial=True)

    def _fill(self, hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _font(self, hex_color="CCEEEE", bold=False, size=10, name="Arial"):
        return Font(name=name, color=hex_color, bold=bold, size=size)

    def _border(self):
        s = Side(style="thin", color="1A3A5A")
        return Border(left=s, right=s, top=s, bottom=s)

    def _center(self, wrap=False):
        return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

    def _apply_header_row(self, ws, row, values, bg, fg, bold=True, size=10, height=22):
        ws.row_dimensions[row].height = height
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font      = self._font(fg, bold=bold, size=size)
            cell.fill      = self._fill(bg)
            cell.alignment = self._center(wrap=True)
            cell.border    = self._border()

    def _setup_data_sheet(self):
        ws = self.ws_data
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A3"

        ws.merge_cells("A1:H1")
        title = ws["A1"]
        title.value     = "ESP32 · MQ-135  —  Promedio cada 10 s"
        title.font      = self._font(self.COLOR_HEADER_FG, bold=True, size=13)
        title.fill      = self._fill(self.COLOR_HEADER_BG)
        title.alignment = self._center()
        title.border    = self._border()
        ws.row_dimensions[1].height = 30

        headers = [
            "N°", "Fecha", "Hora", "ADC Promedio",
            "Voltaje Prom.", "Escala Prom.", "Calidad Promedio", "Sesión"
        ]
        widths  = [6, 14, 11, 14, 12, 11, 16, 22]
        self._apply_header_row(ws, 2, headers,
                               self.COLOR_SUBHDR_BG, self.COLOR_SUBHDR_FG,
                               bold=True, size=10, height=20)
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _write_data_row(self, ws, row, n, dt, val, label, session_name):
        row_bg = self.COLOR_ROW_EVEN if row % 2 == 0 else self.COLOR_ROW_ODD
        quality_color = self.CALIDAD_COLORS.get(label, "CCEEEE")
        voltage = round(val * VREF / ADC_MAX, 4)
        pct     = round(val / ADC_MAX * 100, 2)

        data = [
            n,
            dt.strftime("%Y-%m-%d"),
            dt.strftime("%H:%M:%S.") + f"{dt.microsecond // 1000:03d}",
            val,
            voltage,
            pct,
            label,
            session_name,
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.fill   = self._fill(row_bg)
            cell.border = self._border()
            cell.font   = self._font("CCEEEE", size=9)
            cell.alignment = self._center()

        qa_cell = ws.cell(row=row, column=7)
        qa_cell.font = self._font(quality_color, bold=True, size=9)

        adc_cell = ws.cell(row=row, column=4)
        adc_cell.font = self._font(quality_color, bold=True, size=9)

    def _setup_stats_sheet(self):
        ws = self.ws_stats
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:D1")
        t = ws["A1"]
        t.value     = "Estadísticas de Sesión · MQ-135"
        t.font      = self._font(self.COLOR_HEADER_FG, bold=True, size=13)
        t.fill      = self._fill(self.COLOR_HEADER_BG)
        t.alignment = self._center()
        t.border    = self._border()
        ws.row_dimensions[1].height = 30

        headers = ["Métrica", "Valor", "Unidad", "Descripción"]
        self._apply_header_row(ws, 2, headers,
                               self.COLOR_SUBHDR_BG, self.COLOR_SUBHDR_FG,
                               height=20)
        widths = [22, 16, 12, 36]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        metrics = [
            ("Total de promedios",      "=COUNTA(Lecturas!D3:D100000)-1", "",    "Ventanas de 10 s registradas"),
            ("ADC Prom. Mínimo",        "=IF(COUNTA(Lecturas!D3:D100000)>1,MIN(Lecturas!D3:D100000),\"—\")", "ADC", "Mínimo de los promedios"),
            ("ADC Prom. Máximo",        "=IF(COUNTA(Lecturas!D3:D100000)>1,MAX(Lecturas!D3:D100000),\"—\")", "ADC", "Máximo de los promedios"),
            ("ADC Prom. Promedio",      "=IF(COUNTA(Lecturas!D3:D100000)>1,ROUND(AVERAGE(Lecturas!D3:D100000),1),\"—\")", "ADC", "Media de los promedios"),
            ("Desv. Estándar",          "=IF(COUNTA(Lecturas!D3:D100000)>1,ROUND(STDEV(Lecturas!D3:D100000),2),\"—\")", "ADC", "Dispersión de los promedios"),
            ("Voltaje Prom. Mínimo",    "=IF(COUNTA(Lecturas!E3:E100000)>1,MIN(Lecturas!E3:E100000),\"—\")", "V",   ""),
            ("Voltaje Prom. Máximo",    "=IF(COUNTA(Lecturas!E3:E100000)>1,MAX(Lecturas!E3:E100000),\"—\")", "V",   ""),
            ("Voltaje Prom. Promedio",  "=IF(COUNTA(Lecturas!E3:E100000)>1,ROUND(AVERAGE(Lecturas!E3:E100000),4),\"—\")", "V", ""),
            ("Escala Prom. Promedio",   "=IF(COUNTA(Lecturas!F3:F100000)>1,ROUND(AVERAGE(Lecturas!F3:F100000),2),\"—\")", "%",  ""),
            ("Primer promedio",          "=IF(COUNTA(Lecturas!B3:B3)>0,Lecturas!B3&\" \"&Lecturas!C3,\"—\")", "",   "Fecha y hora del primer promedio"),
            ("Último promedio",         "=IF(COUNTA(Lecturas!D3:D100000)>1,INDEX(Lecturas!B3:B100000,COUNTA(Lecturas!B3:B100000))&\" \"&INDEX(Lecturas!C3:C100000,COUNTA(Lecturas!C3:C100000)),\"—\")", "", "Fecha y hora del último promedio"),
        ]
        for i, (metric, formula, unit, desc) in enumerate(metrics, 3):
            ws.row_dimensions[i].height = 18
            for col, val in enumerate([metric, formula, unit, desc], 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.fill      = self._fill(self.COLOR_STATS_BG)
                cell.border    = self._border()
                cell.alignment = self._center()
                cell.font      = self._font("CCEEEE", size=10, bold=(col == 1))

        row = len(metrics) + 4
        ws.merge_cells(f"A{row}:D{row}")
        h = ws.cell(row=row, column=1, value="Distribución por Calidad del Aire (promedios)")
        h.font      = self._font(self.COLOR_HEADER_FG, bold=True, size=11)
        h.fill      = self._fill(self.COLOR_SUBHDR_BG)
        h.alignment = self._center()
        h.border    = self._border()
        ws.row_dimensions[row].height = 22

        row += 1
        for col, hdr in enumerate(["Categoría", "Lecturas", "%", "Color referencia"], 1):
            c = ws.cell(row=row, column=col, value=hdr)
            c.font      = self._font(self.COLOR_SUBHDR_FG, bold=True, size=9)
            c.fill      = self._fill(self.COLOR_SUBHDR_BG)
            c.alignment = self._center()
            c.border    = self._border()

        for q_label, q_color in self.CALIDAD_COLORS.items():
            row += 1
            ws.row_dimensions[row].height = 16
            cat_formula = f'=COUNTIF(Lecturas!G3:G100000,"{q_label}")'
            pct_formula = f"=IF(B{row}>0,ROUND(B{row}/B3*100,1),0)"
            for col, val in enumerate([q_label, cat_formula, pct_formula, ""], 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill      = self._fill(self.COLOR_STATS_BG)
                cell.border    = self._border()
                cell.alignment = self._center()
                cell.font      = self._font(q_color if col in (1, 2) else "CCEEEE",
                                           bold=(col == 1), size=9)
            color_cell = ws.cell(row=row, column=4, value="█████")
            color_cell.font = self._font(q_color, bold=True, size=12)

    def _setup_info_sheet(self):
        ws = self.ws_info
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:B1")
        t = ws["A1"]
        t.value     = "Información del Sistema"
        t.font      = self._font(self.COLOR_HEADER_FG, bold=True, size=13)
        t.fill      = self._fill(self.COLOR_HEADER_BG)
        t.alignment = self._center()
        t.border    = self._border()
        ws.row_dimensions[1].height = 30

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 38

        info_rows = [
            ("Sensor",              "MQ-135 (Gas: CO₂, NH₃, Benceno, Humo)"),
            ("Microcontrolador",    "ESP32"),
            ("Resolución ADC",      "12-bit (0 – 4095)"),
            ("Voltaje de referencia", f"{VREF} V"),
            ("Modo de registro",    "Promedio cada 10 segundos en Excel"),
            ("Fecha de inicio",     datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Versión dashboard",   "2.0 · ESP32 MQ-135 Pro"),
            ("Archivo generado",    os.path.basename(self.filepath)),
        ]
        for i, (key, val) in enumerate(info_rows, 2):
            ws.row_dimensions[i].height = 18
            bg = self.COLOR_ROW_EVEN if i % 2 == 0 else self.COLOR_ROW_ODD
            for col, v in enumerate([key, val], 1):
                cell = ws.cell(row=i, column=col, value=v)
                cell.fill      = self._fill(bg)
                cell.border    = self._border()
                cell.alignment = self._center()
                cell.font      = self._font("CCEEEE", bold=(col == 1), size=10)

        row = len(info_rows) + 3
        ws.merge_cells(f"A{row}:B{row}")
        h = ws.cell(row=row, column=1, value="Tabla de Referencia · Calidad del Aire")
        h.font      = self._font(self.COLOR_HEADER_FG, bold=True, size=11)
        h.fill      = self._fill(self.COLOR_SUBHDR_BG)
        h.alignment = self._center()
        h.border    = self._border()
        ws.row_dimensions[row].height = 20

        for lo, hi, label, _ in AIR_QUALITY:
            row += 1
            q_color = self.CALIDAD_COLORS.get(label, "CCEEEE")
            ws.row_dimensions[row].height = 16
            for col, val in enumerate([f"{lo} – {hi}  ADC", label], 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill      = self._fill(self.COLOR_STATS_BG)
                cell.border    = self._border()
                cell.alignment = self._center()
                cell.font      = self._font(q_color, bold=(col == 2), size=10)

    def log(self, val: int, label: str, session_name: str):
        """Escribe una fila en el Excel (usado para promedios de 10 s)."""
        with self._lock:
            self._count += 1
            dt = datetime.datetime.now()
            self._write_data_row(
                self.ws_data, self._row + 1,
                self._count, dt, val, label, session_name
            )
            self._row += 1

    def _save(self, initial=False):
        with self._lock:
            dir_name = os.path.dirname(self.filepath)
            base_name = os.path.basename(self.filepath)
            try:
                with tempfile.NamedTemporaryFile(
                    dir=dir_name, prefix="~$", suffix=".xlsx", delete=False
                ) as tmp:
                    tmp_name = tmp.name
                self.wb.save(tmp_name)
                if os.path.exists(self.filepath):
                    os.replace(tmp_name, self.filepath)
                else:
                    os.rename(tmp_name, self.filepath)
                return True
            except Exception as e:
                self.log(f"❌ Error al guardar Excel: {e}", "danger")
                if 'tmp_name' in locals() and os.path.exists(tmp_name):
                    try:
                        os.remove(tmp_name)
                    except:
                        pass
                if initial:
                    self.log("⚠️ El archivo Excel NO se pudo crear. Verifica permisos o ruta.", "danger")
                return False

    def close(self):
        ok = self._save()
        with self._lock:
            if ok:
                self.log("Excel guardado correctamente al cerrar", "ok")
            else:
                self.log("No se pudo guardar el Excel final (el archivo puede estar incompleto)", "danger")

    def export_now(self, dest_path: str) -> bool:
        with self._lock:
            try:
                dir_name = os.path.dirname(dest_path)
                with tempfile.NamedTemporaryFile(
                    dir=dir_name, prefix="~$", suffix=".xlsx", delete=False
                ) as tmp:
                    tmp_name = tmp.name
                self.wb.save(tmp_name)
                if os.path.exists(dest_path):
                    os.replace(tmp_name, dest_path)
                else:
                    os.rename(tmp_name, dest_path)
                return True
            except Exception as e:
                self.log(f"Error al exportar copia: {e}", "danger")
                return False


# ══════════════════════════════════════════════
#  WIDGET: Manómetro Canvas
# ══════════════════════════════════════════════
class Manometer(tk.Canvas):
    def __init__(self, parent, size=260, **kw):
        super().__init__(parent, width=size, height=size // 2 + 30,
                         bg=C["panel"], highlightthickness=0, **kw)
        self.size   = size
        self.cx     = size // 2
        self.cy     = size // 2 + 10
        self.r_arc  = size // 2 - 20
        self._angle = 180.0
        self._draw_static()
        self._needle_id = None
        self._hub_id    = None
        self._draw_needle(180.0, C["hud"])

    def _draw_static(self):
        cx, cy, r = self.cx, self.cy, self.r_arc
        self.create_arc(cx - r, cy - r, cx + r, cy + r,
                        start=0, extent=180,
                        style=tk.ARC, outline=C["border"], width=18)
        colors = ["#00ff88","#aaff00","#ffaa00","#ff6600","#ff3333"]
        seg = 180 / len(colors)
        for i, col in enumerate(colors):
            self.create_arc(cx - r, cy - r, cx + r, cy + r,
                            start=i * seg, extent=seg,
                            style=tk.ARC, outline=col, width=18)
        self.create_rectangle(0, cy, self.size, cy + 50,
                              fill=C["panel"], outline="")
        for i in range(11):
            a = math.radians(180 - i * 18)
            r_outer = r - 9
            r_inner = r - 18 if i % 5 == 0 else r - 13
            x1 = cx + r_outer * math.cos(a)
            y1 = cy - r_outer * math.sin(a)
            x2 = cx + r_inner * math.cos(a)
            y2 = cy - r_inner * math.sin(a)
            w  = 2 if i % 5 == 0 else 1
            self.create_line(x1, y1, x2, y2, fill=C["border"], width=w)
        for angle_d, label in [(180, "0"), (90, "2048"), (0, "4095")]:
            a = math.radians(angle_d)
            lx = cx + (r - 30) * math.cos(a)
            ly = cy - (r - 30) * math.sin(a)
            self.create_text(lx, ly, text=label,
                             fill=C["text"], font=("Courier", 8))

    def _draw_needle(self, angle_deg, color):
        cx, cy = self.cx, self.cy
        if self._needle_id:
            self.delete(self._needle_id)
        if self._hub_id:
            self.delete(self._hub_id)
        tx, ty = needle_coords(cx, cy, self.r_arc - 10, angle_deg)
        self.create_line(cx, cy, tx, ty,
                         fill="#000000", width=5,
                         capstyle=tk.ROUND, tags="shadow")
        self._needle_id = self.create_line(cx, cy, tx, ty,
                                           fill=color, width=3,
                                           capstyle=tk.ROUND)
        r_hub = 8
        self._hub_id = self.create_oval(cx - r_hub, cy - r_hub,
                                        cx + r_hub, cy + r_hub,
                                        fill=C["panel"], outline=C["border"], width=1)
        self.create_oval(cx - 4, cy - 4, cx + 4, cy + 4,
                         fill=color, outline="")

    def set_value(self, val):
        angle  = adc_to_angle(val)
        _, col = classify(val)
        self._draw_needle(angle, col)


# ══════════════════════════════════════════════
#  WIDGET: Waveform
# ══════════════════════════════════════════════
class WaveCanvas(tk.Canvas):
    def __init__(self, parent, width=560, height=70, **kw):
        super().__init__(parent, width=width, height=height,
                         bg=C["panel"], highlightthickness=0, **kw)
        self.w = width
        self.h = height
        self._buf = collections.deque(maxlen=HISTORY_LEN)

    def push(self, val):
        self._buf.append(val)
        self._redraw()

    def _redraw(self):
        self.delete("all")
        buf = list(self._buf)
        n   = len(buf)
        if n < 2:
            return
        _, col = classify(buf[-1])
        pts = []
        for i, v in enumerate(buf):
            x = i / (HISTORY_LEN - 1) * self.w
            y = self.h - (v / ADC_MAX) * (self.h - 4) - 2
            pts.append((x, y))
        flat = [c for p in pts for c in p]
        self.create_line(*flat, fill=col, width=1.5, smooth=True)
        poly = flat + [self.w, self.h, 0, self.h]
        self.create_polygon(*poly, fill=col, stipple="gray25", outline="")


# ══════════════════════════════════════════════
#  WIDGET: Barra de calidad
# ══════════════════════════════════════════════
class AirBar(tk.Canvas):
    def __init__(self, parent, width=560, height=14, **kw):
        super().__init__(parent, width=width, height=height,
                         bg=C["panel"], highlightthickness=0, **kw)
        self.w = width
        self.h = height
        self._draw_bar()
        self._needle = None

    def _draw_bar(self):
        colors = ["#00ff88","#aaff00","#ffaa00","#ff6600","#ff3333"]
        seg_w  = self.w / len(colors)
        for i, c in enumerate(colors):
            self.create_rectangle(i * seg_w, 0,
                                  (i + 1) * seg_w, self.h,
                                  fill=c, outline="")
        self.create_rectangle(0, 0, self.w, self.h,
                              fill="", outline=C["border"], width=1)

    def set_value(self, val):
        if self._needle:
            self.delete(self._needle)
        x = val / ADC_MAX * self.w
        self._needle = self.create_rectangle(
            x - 2, -2, x + 2, self.h + 2,
            fill="white", outline="")


# ══════════════════════════════════════════════
#  HILO SERIAL
# ══════════════════════════════════════════════
class SerialReader(threading.Thread):
    def __init__(self, port, baud, callback_data, callback_err):
        super().__init__(daemon=True)
        self.port     = port
        self.baud     = baud
        self.on_data  = callback_data
        self.on_err   = callback_err
        self._running = True
        self._ser     = None

    def run(self):
        try:
            self._ser = serial.Serial(self.port, self.baud, timeout=1)
            self._ser.flushInput()
            while self._running:
                if self._ser.in_waiting > 0:
                    try:
                        raw = self._ser.readline().decode("utf-8", errors="ignore").strip()
                        if raw:
                            val = int(float(raw))
                            self.on_data(val)
                    except ValueError:
                        pass
                    except Exception as e:
                        self.on_err(str(e))
                time.sleep(0.01)
        except Exception as e:
            self.on_err(str(e))

    def stop(self):
        self._running = False
        if self._ser and self._ser.is_open:
            self._ser.close()


# ══════════════════════════════════════════════
#  HILO SIMULACIÓN
# ══════════════════════════════════════════════
class SimReader(threading.Thread):
    def __init__(self, callback_data):
        super().__init__(daemon=True)
        self.on_data  = callback_data
        self._running = True
        self._phase   = 0.0

    def run(self):
        while self._running:
            self._phase += 0.04
            val = int(2000 + math.sin(self._phase) * 1400
                      + random.gauss(0, 60))
            val = max(0, min(ADC_MAX, val))
            self.on_data(val)
            time.sleep(0.08)

    def stop(self):
        self._running = False


# ══════════════════════════════════════════════
#  VENTANA PRINCIPAL
# ══════════════════════════════════════════════
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("ESP32 · MQ-135 Dashboard")
        self.configure(bg=C["bg"])
        self.resizable(False, False)

        # ── Ventana de 10 segundos para el promedio del Excel ──
        self._window_period = 10.0
        self._window_start  = time.time()
        self._window_vals   = []          # lista de valores ADC crudos

        # ── Excel logger ──
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            f"MQ135_Sesion_{ts}.xlsx"
        )
        self._session_name = f"Sesión {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
        self._excel = ExcelLogger(excel_path, logger_callback=self._log_write)
        self._excel_path = excel_path

        # ── CSV en tiempo real ──
        csv_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "MQ135_Live.csv"
        )
        self._csv = CsvLogger(csv_path)
        self._csv_path = csv_path

        self._reader  = None
        self._sim     = None
        self._running = False
        self._sim_on  = False
        self._stats   = {"min": None, "max": None, "sum": 0, "n": 0}
        self._data_q  = collections.deque()
        self._log_q   = collections.deque()

        self._build_ui()
        self._after_id = None
        self._schedule_update()
        self._start_sim()

    def _build_ui(self):
        # ── HEADER ──
        hdr = tk.Frame(self, bg=C["bg"])
        hdr.pack(fill=tk.X, padx=12, pady=(10, 4))

        tk.Label(hdr, text="ESP32 · MQ-135 MONITOR",
                 bg=C["bg"], fg=C["hud"],
                 font=("Courier", 11, "bold")).pack(side=tk.LEFT)

        self._dot_var = tk.StringVar(value="●")
        self._dot_lbl = tk.Label(hdr, textvariable=self._dot_var,
                                 bg=C["bg"], fg=C["danger"],
                                 font=("Courier", 14))
        self._dot_lbl.pack(side=tk.RIGHT, padx=(0, 4))

        self._status_var = tk.StringVar(value="DESCONECTADO")
        tk.Label(hdr, textvariable=self._status_var,
                 bg=C["bg"], fg=C["text"],
                 font=("Courier", 9)).pack(side=tk.RIGHT, padx=4)

        self._excel_var = tk.StringVar(
            value=f"📊 {os.path.basename(self._excel_path)}")
        tk.Label(hdr, textvariable=self._excel_var,
                 bg=C["bg"], fg=C["amber"],
                 font=("Courier", 8)).pack(side=tk.RIGHT, padx=8)

        tk.Frame(self, bg=C["border"], height=1).pack(fill=tk.X, padx=12)

        # ── BODY ──
        body = tk.Frame(self, bg=C["bg"])
        body.pack(fill=tk.BOTH, padx=12, pady=6)

        left  = tk.Frame(body, bg=C["bg"])
        right = tk.Frame(body, bg=C["bg"])
        left.pack(side=tk.LEFT, fill=tk.BOTH)
        right.pack(side=tk.LEFT, fill=tk.BOTH, padx=(12, 0))

        # manómetro
        self._gauge = Manometer(left, size=280)
        self._gauge.pack()

        self._adc_var = tk.StringVar(value="—")
        self._adc_lbl = tk.Label(left, textvariable=self._adc_var,
                                 bg=C["panel"], fg=C["hud"],
                                 font=("Courier", 32, "bold"))
        self._adc_lbl.pack(fill=tk.X, ipadx=10, ipady=4)

        self._badge_var = tk.StringVar(value="ESPERANDO")
        self._badge_lbl = tk.Label(left, textvariable=self._badge_var,
                                   bg=C["panel"], fg=C["ok"],
                                   font=("Courier", 11, "bold"),
                                   relief="flat")
        self._badge_lbl.pack(fill=tk.X, ipady=3)

        cards = tk.Frame(left, bg=C["bg"])
        cards.pack(fill=tk.X, pady=6)

        self._volt_var = tk.StringVar(value="—")
        self._pct_var  = tk.StringVar(value="—")
        self._time_var = tk.StringVar(value="00:00:00")

        for label, var, unit in [
            ("VOLTAJE",  self._volt_var, "V"),
            ("ESCALA",   self._pct_var,  "%"),
            ("TIEMPO",   self._time_var, ""),
        ]:
            f = tk.Frame(cards, bg=C["panel"], relief="flat", bd=0,
                         highlightbackground=C["border"], highlightthickness=1)
            f.pack(side=tk.LEFT, expand=True, fill=tk.X,
                   padx=3, ipadx=6, ipady=6)
            tk.Label(f, text=label, bg=C["panel"], fg=C["text"],
                     font=("Courier", 7)).pack()
            vf = tk.Frame(f, bg=C["panel"])
            vf.pack()
            tk.Label(vf, textvariable=var, bg=C["panel"],
                     fg=C["bright"], font=("Courier", 14, "bold")).pack(side=tk.LEFT)
            if unit:
                tk.Label(vf, text=unit, bg=C["panel"],
                         fg=C["text"], font=("Courier", 9)).pack(side=tk.LEFT, pady=(4,0))

        # ── DERECHA ──
        # conexión serial
        conn_frm = tk.LabelFrame(right, text=" CONEXIÓN SERIAL ",
                                  bg=C["panel"], fg=C["text"],
                                  font=("Courier", 8), bd=1, relief="flat",
                                  highlightbackground=C["border"],
                                  highlightthickness=1)
        conn_frm.pack(fill=tk.X, pady=(0, 6))

        row1 = tk.Frame(conn_frm, bg=C["panel"])
        row1.pack(fill=tk.X, padx=6, pady=4)

        tk.Label(row1, text="Puerto:", bg=C["panel"],
                 fg=C["text"], font=("Courier", 9)).pack(side=tk.LEFT)
        self._port_var = tk.StringVar(value=DEFAULT_PORT)
        tk.Entry(row1, textvariable=self._port_var, width=10,
                 bg=C["bg"], fg=C["bright"], insertbackground=C["hud"],
                 font=("Courier", 10), relief="flat",
                 highlightbackground=C["border"], highlightthickness=1
                 ).pack(side=tk.LEFT, padx=4)

        tk.Label(row1, text="Baud:", bg=C["panel"],
                 fg=C["text"], font=("Courier", 9)).pack(side=tk.LEFT)
        self._baud_var = tk.StringVar(value=str(DEFAULT_BAUD))
        baud_cb = ttk.Combobox(row1, textvariable=self._baud_var,
                               values=["9600","57600","115200","230400"],
                               width=7, state="readonly",
                               font=("Courier", 9))
        baud_cb.pack(side=tk.LEFT, padx=4)

        row2 = tk.Frame(conn_frm, bg=C["panel"])
        row2.pack(fill=tk.X, padx=6, pady=(0, 4))

        self._conn_btn = tk.Button(
            row2, text="CONECTAR",
            bg=C["bg"], fg=C["ok"],
            font=("Courier", 9, "bold"),
            activebackground=C["panel"], activeforeground=C["ok"],
            relief="flat",
            highlightbackground=C["ok"], highlightthickness=1,
            command=self._toggle_connect, padx=10, pady=4)
        self._conn_btn.pack(side=tk.LEFT, padx=(0, 6))

        self._sim_btn = tk.Button(
            row2, text="▶ SIMULACIÓN",
            bg=C["bg"], fg=C["hud"],
            font=("Courier", 9, "bold"),
            activebackground=C["panel"], activeforeground=C["hud"],
            relief="flat",
            highlightbackground=C["hud"], highlightthickness=1,
            command=self._toggle_sim, padx=10, pady=4)
        self._sim_btn.pack(side=tk.LEFT)

        row3 = tk.Frame(conn_frm, bg=C["panel"])
        row3.pack(fill=tk.X, padx=6, pady=(0, 4))
        tk.Label(row3, text="Puertos detectados:", bg=C["panel"],
                 fg=C["text"], font=("Courier", 8)).pack(side=tk.LEFT)
        self._ports_var = tk.StringVar(value=self._scan_ports())
        tk.Label(row3, textvariable=self._ports_var,
                 bg=C["panel"], fg=C["bright"],
                 font=("Courier", 8)).pack(side=tk.LEFT, padx=4)
        tk.Button(row3, text="⟳", bg=C["panel"], fg=C["text"],
                  relief="flat", font=("Courier", 9),
                  command=lambda: self._ports_var.set(self._scan_ports())
                  ).pack(side=tk.LEFT)

        # estadísticas
        stats_frm = tk.LabelFrame(right, text=" ESTADÍSTICAS ",
                                   bg=C["panel"], fg=C["text"],
                                   font=("Courier", 8), bd=1, relief="flat",
                                   highlightbackground=C["border"],
                                   highlightthickness=1)
        stats_frm.pack(fill=tk.X, pady=(0, 6))

        srow = tk.Frame(stats_frm, bg=C["panel"])
        srow.pack(fill=tk.X, padx=6, pady=6)

        self._min_var = tk.StringVar(value="—")
        self._avg_var = tk.StringVar(value="—")
        self._max_var = tk.StringVar(value="—")
        self._cnt_var = tk.StringVar(value="0")

        for lbl, var in [("MÍN", self._min_var),
                         ("PROM", self._avg_var),
                         ("MÁX", self._max_var),
                         ("LECTURAS", self._cnt_var)]:
            sf = tk.Frame(srow, bg=C["panel"])
            sf.pack(side=tk.LEFT, expand=True)
            tk.Label(sf, text=lbl, bg=C["panel"], fg=C["text"],
                     font=("Courier", 7)).pack()
            tk.Label(sf, textvariable=var, bg=C["panel"],
                     fg=C["bright"], font=("Courier", 13, "bold")).pack()

        tk.Button(stats_frm, text="⟳ Reset",
                  bg=C["panel"], fg=C["text"],
                  font=("Courier", 8), relief="flat",
                  command=self._reset_stats).pack(anchor="e", padx=6, pady=(0, 4))

        # calidad del aire
        qa_frm = tk.LabelFrame(right, text=" ÍNDICE DE CALIDAD DEL AIRE ",
                                bg=C["panel"], fg=C["text"],
                                font=("Courier", 8), bd=1, relief="flat",
                                highlightbackground=C["border"],
                                highlightthickness=1)
        qa_frm.pack(fill=tk.X, pady=(0, 6))

        self._airbar = AirBar(qa_frm, width=390, height=14)
        self._airbar.pack(padx=8, pady=(6, 2))

        scale_lbl_frm = tk.Frame(qa_frm, bg=C["panel"])
        scale_lbl_frm.pack(fill=tk.X, padx=8, pady=(0, 6))
        for txt in ["Excelente","Bueno","Moderado","Malo","Peligroso"]:
            tk.Label(scale_lbl_frm, text=txt, bg=C["panel"],
                     fg=C["text"], font=("Courier", 7)
                     ).pack(side=tk.LEFT, expand=True)

        # waveform
        wave_frm = tk.LabelFrame(right, text=" HISTORIAL · ÚLTIMAS 80 LECTURAS ",
                                  bg=C["panel"], fg=C["text"],
                                  font=("Courier", 8), bd=1, relief="flat",
                                  highlightbackground=C["border"],
                                  highlightthickness=1)
        wave_frm.pack(fill=tk.X, pady=(0, 6))

        self._wave = WaveCanvas(wave_frm, width=390, height=70)
        self._wave.pack(padx=8, pady=6)

        # log
        log_frm = tk.LabelFrame(right, text=" LOG DEL SISTEMA ",
                                  bg=C["panel"], fg=C["text"],
                                  font=("Courier", 8), bd=1, relief="flat",
                                  highlightbackground=C["border"],
                                  highlightthickness=1)
        log_frm.pack(fill=tk.BOTH, expand=True)

        self._log = tk.Text(log_frm, height=6, width=48,
                            bg=C["bg"], fg=C["text"],
                            font=("Courier", 8), relief="flat",
                            insertbackground=C["hud"],
                            state="disabled")
        self._log.pack(fill=tk.BOTH, expand=True, padx=6, pady=4)
        self._log.tag_config("ok",     foreground=C["ok"])
        self._log.tag_config("warn",   foreground=C["warn"])
        self._log.tag_config("danger", foreground=C["danger"])
        self._log.tag_config("info",   foreground=C["hud"])

        tk.Button(log_frm, text="✕ Limpiar", bg=C["panel"],
                  fg=C["text"], font=("Courier", 8), relief="flat",
                  command=self._clear_log).pack(anchor="e", padx=6, pady=(0, 4))

        # ── BOTÓN EXPORTAR EXCEL ──
        export_frm = tk.Frame(right, bg=C["bg"])
        export_frm.pack(fill=tk.X, pady=(4, 0))

        self._export_btn = tk.Button(
            export_frm,
            text="💾  EXPORTAR EXCEL AHORA",
            bg=C["bg"], fg=C["amber"],
            font=("Courier", 9, "bold"),
            activebackground=C["panel"], activeforeground=C["amber"],
            relief="flat",
            highlightbackground=C["amber"], highlightthickness=1,
            command=self._export_excel_now,
            padx=10, pady=5
        )
        self._export_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 4))

        self._csv_lbl = tk.Label(
            export_frm,
            text="📄 CSV: MQ135_Live.csv  (abrir con Datos→Desde CSV en Excel)",
            bg=C["bg"], fg=C["text"],
            font=("Courier", 7),
            anchor="w"
        )
        self._csv_lbl.pack(side=tk.LEFT)

        self._start_time = time.time()

    # ── Helpers ──
    def _scan_ports(self):
        ports = [p.device for p in serial.tools.list_ports.comports()]
        return "  ".join(ports) if ports else "(ninguno)"

    def _log_write(self, msg, tag=""):
        ts = time.strftime("%H:%M:%S")
        line = f"[{ts}] {msg}\n"
        self._log.config(state="normal")
        self._log.insert(tk.END, line, tag)
        self._log.see(tk.END)
        lines = int(self._log.index("end-1c").split(".")[0])
        if lines > 200:
            self._log.delete("1.0", "50.0")
        self._log.config(state="disabled")

    def _clear_log(self):
        self._log.config(state="normal")
        self._log.delete("1.0", tk.END)
        self._log.config(state="disabled")

    def _reset_stats(self):
        self._stats = {"min": None, "max": None, "sum": 0, "n": 0}
        self._min_var.set("—"); self._avg_var.set("—")
        self._max_var.set("—"); self._cnt_var.set("0")
        self._log_write("Estadísticas reiniciadas", "info")

    def _export_excel_now(self):
        ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            f"MQ135_Export_{ts}.xlsx"
        )
        self._export_btn.config(text="⏳ Exportando...", fg=C["text"])
        self.update_idletasks()

        def _do_export():
            ok = self._excel.export_now(dest)
            def _ui():
                if ok:
                    self._export_btn.config(
                        text="✅ Exportado · abre MQ135_Export_*.xlsx",
                        fg=C["ok"])
                    self._log_write(
                        f"Excel exportado → {os.path.basename(dest)}", "ok")
                else:
                    self._export_btn.config(
                        text="❌ Error al exportar", fg=C["danger"])
                    self._log_write("Error al exportar Excel", "danger")
                self.after(3000, lambda: self._export_btn.config(
                    text="💾  EXPORTAR EXCEL AHORA", fg=C["amber"]))
            self.after(0, _ui)

        threading.Thread(target=_do_export, daemon=True).start()

    # ── Conexión serial ──
    def _toggle_connect(self):
        if self._running:
            self._disconnect()
        else:
            self._connect()

    def _connect(self):
        port = self._port_var.get().strip()
        baud = int(self._baud_var.get())
        self._reader = SerialReader(port, baud, self._on_data, self._on_error)
        self._reader.start()
        self._running = True
        if self._sim_on:
            self._stop_sim()
        self._conn_btn.config(text="DESCONECTAR",
                              fg=C["danger"], highlightbackground=C["danger"])
        self._status_var.set(port.upper())
        self._dot_lbl.config(fg=C["ok"])
        self._log_write(f"Conectado → {port} @ {baud} baud", "ok")

    def _disconnect(self):
        if self._reader:
            self._reader.stop()
            self._reader = None
        self._running = False
        self._conn_btn.config(text="CONECTAR",
                              fg=C["ok"], highlightbackground=C["ok"])
        self._status_var.set("DESCONECTADO")
        self._dot_lbl.config(fg=C["danger"])
        self._log_write("Puerto cerrado", "warn")

    # ── Simulación ──
    def _toggle_sim(self):
        if self._sim_on:
            self._stop_sim()
        else:
            self._start_sim()

    def _start_sim(self):
        if self._running:
            self._disconnect()
        self._sim    = SimReader(self._on_data)
        self._sim_on = True
        self._sim.start()
        self._sim_btn.config(text="⏸ SIMULACIÓN")
        self._status_var.set("SIMULACIÓN")
        self._dot_lbl.config(fg=C["warn"])
        self._log_write("Simulación activa", "info")

    def _stop_sim(self):
        if self._sim:
            self._sim.stop()
            self._sim = None
        self._sim_on = False
        self._sim_btn.config(text="▶ SIMULACIÓN")
        if not self._running:
            self._status_var.set("DESCONECTADO")
            self._dot_lbl.config(fg=C["danger"])
        self._log_write("Simulación pausada", "warn")

    # ── Callbacks ──
    def _on_data(self, val):
        self._data_q.append(val)

    def _on_error(self, msg):
        self._log_q.append(("ERROR: " + msg, "danger"))
        if self._running:
            self._disconnect()

    # ── Loop UI (procesa todas las lecturas, promedia cada 10 s para Excel) ──
    def _schedule_update(self):
        self._process_queue()
        self._after_id = self.after(UPDATE_MS, self._schedule_update)

    def _process_queue(self):
        """Procesa todas las lecturas disponibles, actualiza GUI con la última
        y va acumulando para el promedio de 10 s del Excel."""
        # Recoger todos los valores pendientes
        batch = []
        while self._data_q:
            batch.append(self._data_q.popleft())

        if not batch:
            # Si no hay datos, igual comprobamos la ventana del Excel
            self._check_excel_window()
            # Actualizar reloj de todas formas
            self._update_clock()
            # Procesar log
            self._flush_log()
            return

        # Procesar cada lectura para estadísticas, CSV y acumulación
        for val in batch:
            val = max(0, min(ADC_MAX, val))

            # Estadísticas globales (mín, máx, suma, contador)
            s = self._stats
            if s["min"] is None or val < s["min"]:
                s["min"] = val
            if s["max"] is None or val > s["max"]:
                s["max"] = val
            s["sum"] += val
            s["n"]   += 1

            # Agregar a la ventana de 10 segundos
            self._window_vals.append(val)

            # Guardar en CSV (tiempo real, todas las lecturas)
            label, _ = classify(val)
            self._csv.log(val, label, self._session_name)

        # Actualizar la interfaz con el último valor recibido
        last_val = batch[-1]
        self._update_display(last_val)

        # Verificar si ya se cumplió la ventana de 10 s para el Excel
        self._check_excel_window()

        # Actualizar reloj y log
        self._update_clock()
        self._flush_log()

    def _check_excel_window(self):
        """Si pasaron >= 10 s desde el inicio de la ventana, guarda el promedio en Excel."""
        now = time.time()
        if now - self._window_start >= self._window_period:
            if self._window_vals:   # solo si hay datos en la ventana
                avg_val = round(sum(self._window_vals) / len(self._window_vals))
                label, _ = classify(avg_val)
                # Guardar en Excel (se llama en un hilo separado para no bloquear)
                threading.Thread(
                    target=self._excel.log,
                    args=(avg_val, label, self._session_name),
                    daemon=True
                ).start()
                self._log_write(
                    f"📊 Promedio 10s: {avg_val} ADC → {label} (guardado en Excel)",
                    "info"
                )
            # Reiniciar ventana
            self._window_start = now
            self._window_vals.clear()

    def _update_display(self, val):
        """Refresca manómetro, waveform, barra, etiquetas y estadísticas con el valor dado."""
        label, color = classify(val)

        # Manómetro
        self._gauge.set_value(val)
        # Forma de onda
        self._wave.push(val)
        # Barra de calidad
        self._airbar.set_value(val)

        # Valor ADC grande
        self._adc_var.set(str(val))
        self._adc_lbl.config(fg=color)

        # Insignia de calidad
        self._badge_var.set(label)
        self._badge_lbl.config(fg=color)

        # Voltaje y porcentaje
        self._volt_var.set(f"{val * VREF / ADC_MAX:.3f}")
        self._pct_var.set(f"{val / ADC_MAX * 100:.1f}")

        # Estadísticas acumuladas
        s = self._stats
        self._min_var.set(str(s["min"]) if s["min"] is not None else "—")
        self._max_var.set(str(s["max"]) if s["max"] is not None else "—")
        if s["n"] > 0:
            self._avg_var.set(str(round(s["sum"] / s["n"])))
        else:
            self._avg_var.set("—")
        self._cnt_var.set(str(s["n"]))

    def _update_clock(self):
        elapsed = int(time.time() - self._start_time)
        h, rem  = divmod(elapsed, 3600)
        m, s    = divmod(rem, 60)
        self._time_var.set(f"{h:02d}:{m:02d}:{s:02d}")

    def _flush_log(self):
        while self._log_q:
            msg, tag = self._log_q.popleft()
            self._log_write(msg, tag)

    def on_close(self):
        self._disconnect()
        self._stop_sim()
        if self._after_id:
            self.after_cancel(self._after_id)
        # Guardar cualquier promedio pendiente en el Excel
        if self._window_vals:
            avg_val = round(sum(self._window_vals) / len(self._window_vals))
            label, _ = classify(avg_val)
            self._excel.log(avg_val, label, self._session_name)
            self._window_vals.clear()
        self._excel.close()
        self._csv.close()
        self._log_write("Excel y CSV guardados y cerrados", "ok")
        self.destroy()


if __name__ == "__main__":
    app = App()
    app.protocol("WM_DELETE_WINDOW", app.on_close)
    app.mainloop()